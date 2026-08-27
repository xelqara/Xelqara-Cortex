"""Safe, local document adapters for BidCore.

The adapters extract text and structural metadata without executing macros,
embedded scripts, or external links. XLSX support uses openpyxl when present;
DOCX is parsed from its XML package; PDF uses the installed pdftotext binary.
"""
from __future__ import annotations

import csv
import re
import shutil
import subprocess
import tempfile
import zipfile
from dataclasses import dataclass, asdict
from pathlib import Path
from xml.etree import ElementTree

@dataclass(frozen=True)
class DocumentChunk:
    source: str
    location: str
    text: str
    metadata: dict

class DocumentImportError(RuntimeError):
    pass

class DocumentImporter:
    def import_file(self, path: str | Path) -> list[DocumentChunk]:
        file = Path(path).expanduser().resolve()
        if not file.is_file():
            raise FileNotFoundError(file)
        suffix = file.suffix.lower()
        if suffix in {".csv", ".tsv"}:
            return self._csv(file)
        if suffix == ".xlsx":
            return self._xlsx(file)
        if suffix == ".docx":
            return self._docx(file)
        if suffix == ".pdf":
            return self._pdf(file)
        if suffix in {".md", ".txt"}:
            text = file.read_text(encoding="utf-8", errors="replace")
            return [DocumentChunk(file.name, "text", text, {"format": suffix[1:]})]
        raise DocumentImportError(f"unsupported document format: {suffix}")

    @staticmethod
    def _csv(file: Path) -> list[DocumentChunk]:
        delimiter = "\t" if file.suffix.lower() == ".tsv" else ","
        chunks = []
        with file.open(newline="", encoding="utf-8-sig", errors="replace") as stream:
            for row_number, row in enumerate(csv.reader(stream, delimiter=delimiter), 1):
                text = " | ".join(cell.strip() for cell in row if cell and cell.strip())
                if text:
                    chunks.append(DocumentChunk(file.name, f"row:{row_number}", text, {"format": "csv", "row": row_number}))
        return chunks

    @staticmethod
    def _xlsx(file: Path) -> list[DocumentChunk]:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise DocumentImportError("XLSX support requires optional dependency openpyxl") from exc
        chunks = []
        workbook = load_workbook(file, read_only=True, data_only=True, keep_links=False)
        for sheet in workbook.worksheets:
            for row_number, row in enumerate(sheet.iter_rows(values_only=True), 1):
                cells = [str(value).strip() for value in row if value is not None and str(value).strip()]
                if cells:
                    text = " | ".join(cells)
                    chunks.append(DocumentChunk(file.name, f"sheet:{sheet.title}/row:{row_number}", text, {"format": "xlsx", "sheet": sheet.title, "row": row_number}))
        return chunks

    @staticmethod
    def _docx(file: Path) -> list[DocumentChunk]:
        chunks = []
        try:
            with zipfile.ZipFile(file) as package:
                xml = package.read("word/document.xml")
        except (KeyError, zipfile.BadZipFile) as exc:
            raise DocumentImportError("invalid DOCX package") from exc
        root = ElementTree.fromstring(xml)
        ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
        for index, paragraph in enumerate(root.findall(".//w:p", ns), 1):
            text = "".join(node.text or "" for node in paragraph.findall(".//w:t", ns)).strip()
            if text:
                chunks.append(DocumentChunk(file.name, f"paragraph:{index}", text, {"format": "docx", "paragraph": index}))
        return chunks

    @staticmethod
    def _pdf(file: Path) -> list[DocumentChunk]:
        pdftotext = shutil.which("pdftotext")
        if not pdftotext:
            raise DocumentImportError("PDF support requires the pdftotext utility")
        with tempfile.TemporaryDirectory() as tmp:
            output = Path(tmp) / "text.txt"
            completed = subprocess.run([pdftotext, "-layout", str(file), str(output)], capture_output=True, text=True, timeout=60)
            if completed.returncode != 0:
                raise DocumentImportError(completed.stderr.strip() or "pdftotext failed")
            text = output.read_text(encoding="utf-8", errors="replace")
        chunks = []
        for page_number, page in enumerate(re.split(r"\f", text), 1):
            page = page.strip()
            if page:
                chunks.append(DocumentChunk(file.name, f"page:{page_number}", page, {"format": "pdf", "page": page_number}))
        return chunks

    @staticmethod
    def to_json(chunks: list[DocumentChunk]) -> list[dict]:
        return [asdict(chunk) for chunk in chunks]
