"""Preserve-the-original XLSX response writer for BidCore."""
from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Iterable

from .bidcore import BidCore, BidDraft
from .core import Cortex
from .document_io import DocumentImporter

_QUESTION_HEADERS = {"question", "question text", "questionnaire", "requirement", "security question", "rfi question"}
_ANSWER_HEADERS = {"answer", "response", "vendor response", "comments", "supplier response"}

def _norm(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip().casefold()

def _header_kind(value: object, candidates: set[str]) -> bool:
    normalized = _norm(value)
    return normalized in candidates or any(token in normalized for token in candidates if len(token) > 5)

def _question_text(row: Iterable[object]) -> str:
    cells = [_norm(value) for value in row if _norm(value)]
    if not cells:
        return ""
    candidates = [cell for cell in cells if "?" in cell or len(cell.split()) >= 5]
    return max(candidates, key=len) if candidates else max(cells, key=len)

class XlsxResponseError(RuntimeError):
    pass

class XlsxResponseWriter:
    """Write answer columns while leaving existing workbook structure untouched."""

    def write(self, source: str | Path, output: str | Path, drafts: Iterable[BidDraft]) -> dict[str, object]:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise XlsxResponseError("XLSX response writing requires openpyxl") from exc
        source_path = Path(source).expanduser().resolve()
        output_path = Path(output).expanduser().resolve()
        if source_path.suffix.lower() != ".xlsx" or output_path.suffix.lower() != ".xlsx":
            raise XlsxResponseError("source and output must both be .xlsx files")
        if source_path == output_path:
            raise XlsxResponseError("refusing to overwrite the source workbook")
        if not source_path.is_file():
            raise FileNotFoundError(source_path)
        workbook = load_workbook(source_path, read_only=False, data_only=False, keep_links=False)
        draft_map = {_norm(draft.question): draft for draft in drafts}
        matched = 0
        written = 0
        sheets: list[str] = []
        for sheet in workbook.worksheets:
            header_row = None
            question_col = None
            answer_col = None
            for row_number in range(1, min(sheet.max_row, 30) + 1):
                values = [sheet.cell(row_number, col).value for col in range(1, sheet.max_column + 1)]
                q_cols = [col for col, value in enumerate(values, 1) if _header_kind(value, _QUESTION_HEADERS)]
                a_cols = [col for col, value in enumerate(values, 1) if _header_kind(value, _ANSWER_HEADERS)]
                if q_cols:
                    header_row, question_col = row_number, q_cols[0]
                    answer_col = a_cols[0] if a_cols else None
                    break
            if header_row is None or question_col is None:
                continue
            sheets.append(sheet.title)
            if answer_col is None:
                answer_col = sheet.max_column + 1
                sheet.cell(header_row, answer_col).value = "Xelqara Draft"
            source_col = sheet.max_column + 1
            confidence_col = source_col + 1
            review_col = source_col + 2
            sheet.cell(header_row, source_col).value = "Xelqara Sources"
            sheet.cell(header_row, confidence_col).value = "Xelqara Confidence"
            sheet.cell(header_row, review_col).value = "Xelqara Review"
            for row_number in range(header_row + 1, sheet.max_row + 1):
                question = _question_text(sheet.cell(row_number, col).value for col in range(1, sheet.max_column + 1))
                draft = draft_map.get(question)
                if draft is None:
                    continue
                matched += 1
                sheet.cell(row_number, answer_col).value = draft.draft
                sheet.cell(row_number, source_col).value = " | ".join(draft.sources)
                sheet.cell(row_number, confidence_col).value = draft.confidence
                sheet.cell(row_number, review_col).value = draft.review
                written += 1
        if not sheets:
            raise XlsxResponseError("no worksheet with a recognizable question header was found")
        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)
        return {"source": str(source_path), "output": str(output_path), "sheets": sheets, "drafts": len(draft_map), "matched": matched, "written": written, "policy": "Drafts require human approval before submission."}

def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(prog="bidcore-fill-xlsx", description="Fill a copy of an XLSX questionnaire with BidCore drafts")
    parser.add_argument("input", help="source .xlsx questionnaire")
    parser.add_argument("output", help="new .xlsx output path; source is never overwritten")
    parser.add_argument("--root", default=".cortex")
    parser.add_argument("--limit", type=int, default=3)
    args = parser.parse_args(argv)
    source = Path(args.input)
    cortex = Cortex(args.root)
    try:
        chunks = DocumentImporter().import_file(source)
        questions = BidCore.parse_document_chunks(chunks)
        drafts = BidCore(cortex).draft_batch(questions, max(1, min(args.limit, 5)))
        result = XlsxResponseWriter().write(source, args.output, drafts)
        print(json.dumps(result, ensure_ascii=False, indent=2))
        return 0
    finally:
        cortex.close()

if __name__ == "__main__":
    raise SystemExit(main())
