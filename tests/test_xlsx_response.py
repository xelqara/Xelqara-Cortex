import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from xelqara_cortex import Cortex
from xelqara_cortex.bidcore import BidCore
from xelqara_cortex.document_io import DocumentImporter
from xelqara_cortex.xlsx_response import XlsxResponseWriter


class XlsxResponseTests(unittest.TestCase):
    def test_preserves_source_and_adds_review_columns(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "questionnaire.xlsx"
            output = root / "answered.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Security"
            sheet.append(["ID", "Question", "Response", "Owner"])
            sheet.append(["SEC-1", "Do you encrypt customer data at rest?", "", "Security"])
            sheet.append(["SEC-2", "What is your disaster recovery RTO?", "", "BCP"])
            workbook.save(source)
            cortex = Cortex(root / ".cortex")
            try:
                cortex.ingest_text("security.md", "Customer data is encrypted at rest.")
                questions = BidCore.parse_document_chunks(DocumentImporter().import_file(source))
                drafts = BidCore(cortex).draft_batch(questions)
                result = XlsxResponseWriter().write(source, output, drafts)
                self.assertEqual(result["written"], 2)
                self.assertTrue(source.exists())
                check = load_workbook(output, data_only=False)
                sheet = check["Security"]
                headers = [sheet.cell(1, col).value for col in range(1, sheet.max_column + 1)]
                self.assertIn("Xelqara Sources", headers)
                self.assertIn("Xelqara Review", headers)
                self.assertTrue(sheet.cell(2, 3).value)
                self.assertEqual(sheet.cell(2, 5).value, "security.md")
                self.assertEqual(sheet.cell(2, 7).value, "pending_review")
            finally:
                cortex.close()


if __name__ == "__main__":
    unittest.main()
